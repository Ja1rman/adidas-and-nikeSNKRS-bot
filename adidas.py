# -*- coding: utf-8 -*-

import pause
import random
import re
import threading
import time
import traceback
from datetime import datetime
from os import path

import discord_webhook
import httplib2
import requests
import apiclient
from bs4 import BeautifulSoup
from oauth2client.service_account import ServiceAccountCredentials
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver import Firefox
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.firefox.options import Options
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
import gspread


def filePath():
    return path.dirname(path.abspath(__file__))

counterrrr =  0
wb = load_workbook(filePath() + '/names.xlsx')
sheet = wb['info']


botnumber = str(sheet.cell(row=8, column=1).value)
def ds_sad(asd,strochka):
    discord_webhook.DiscordWebhook(url=str(sheet.cell(row=14,column=1).value), content='Нужен код для ' + str(sheet.cell(row=8, column=1).value) + ' ' + str(asd) + ' вкладки. ' + 'были использованы '  + str(strochka) + ' данные.').execute()
def ds_sad1(asd,strochka):
    discord_webhook.DiscordWebhook(url=str(sheet.cell(row=14,column=1).value), content='Ты ебаный бездарь, поэтому снова нужен код для ' + str(sheet.cell(row=8, column=1).value) + ' ' + str(asd) + ' вкладки ' + str(strochka)).execute()




class adidas:
    def __init__(self):
        #sheesh
        pass

    def main(self, hour, min, sec, j, sheet):
        strokazapolnenia = random.randint(2, int((sheet.cell(row=12, column=1)).value) + 1)
        firstName1 = (str(sheet.cell(row=strokazapolnenia, column=5).value))
        lastName1 = (str(sheet.cell(row=strokazapolnenia, column=6).value))
        city1 = (str(sheet.cell(row=strokazapolnenia, column=7).value))
        zipcode1 = (str(sheet.cell(row=strokazapolnenia, column=8).value))
        address11 =(str(sheet.cell(row=strokazapolnenia, column=9).value))
        houseNumber1 =(str(sheet.cell(row=strokazapolnenia, column=10).value))
        phoneNumber1 = (str(sheet.cell(row=strokazapolnenia, column=11).value))
        emailAddress1 = (str(sheet.cell(row=strokazapolnenia, column=12).value))
        global select_box
        fullprox = str(sheet.cell(row=(j + 1), column=18).value)
        options = Options()
        options.headless = False
        if fullprox != 'None':
            #print(fullprox)
            try:
                ip, port, login, password = fullprox.split(":")
            except:
                ip, port = fullprox.split(":")
            options = webdriver.firefox.options.Options()
            options.add_argument('--disable-blink-features=AutomationControlled')
            options.headless = False
            profile = webdriver.FirefoxProfile()
            profile.set_preference("general.useragent.override",
                                   "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:87.0) Gecko/20100101 Firefox/87.0")

            firefox_capabilities = webdriver.DesiredCapabilities.FIREFOX
            firefox_capabilities["marionette"] = True

            prox = ip + ':' + port
            firefox_capabilities["proxy"] = {
                "proxyType": "MANUAL",
                "httpProxy": prox,
                "ftpProxy": prox,
                "sslProxy": prox
            }

            self.driver = webdriver.Firefox(executable_path=filePath() + '/geckodriver.exe', capabilities=firefox_capabilities,
                                       options=options)
        else:
            self.driver = Firefox(options=options, executable_path=filePath() + '/geckodriver.exe')

        today = datetime.now()
        pause.until(datetime(int(today.year), int(today.month), int(today.day), int(hour), int(min), int(sec), 0))

        while True:
            try:
                global counterrrr
                abs = random.randint(1,7)
                if(abs <= 5):
                    url = 'https://www.adidas.ru/yeezy/product/GV7903'
                else:
                    url = 'https://www.adidas.ru/yeezy/product/FY4567'
                self.driver.get(url)
                pause.until(datetime(2021,5,29,0, 25, 0))
                self.driver.set_window_size(1920, 1080)
                #time.sleep(1000)
                while True:
                    while (counterrrr > 0):
                        time.sleep(30)
                    try:
                        html = self.driver.page_source
                        if "Поздравляем! Ты почти у цели." in html:
                            break
                        else:
                            time.sleep(5)
                    except:
                        time.sleep(5)
                while (counterrrr > 0):
                    time.sleep(30)
                counterrrr += 1
                #randomcolor = random.randint(1, 3)
                '''if (self.driver.current_url == url):
                    try:
                        select_color = self.driver.find_element_by_xpath("//img[@data-auto-id='product-GZ5554']")
                        select_color.click()
                    except:
                        pass
                    try:
                        select_color = self.driver.find_element_by_xpath("//img[@data-auto-id='product-GZ5554-01']")
                        select_color.click()
                    except:
                        pass
                else:
                    try:
                        elem = self.driver.find_element_by_xpath(
                            '//div[@class="src-components-___product__thumbnailContainer___nWuVo"]')
                        child = elem.find_elements_by_xpath(".//*")
                        child[random.randint(0, len(child) - 1)].click()
                    except:
                        randomcolor = random.randint(1, 3)
                        if (randomcolor <= 2):
                            try:
                                select_color = self.driver.find_element_by_xpath("//img[@data-auto-id='product-GZ5554']")
                                select_color.click()
                            except:
                                pass
                            try:
                                select_color = self.driver.find_element_by_xpath("//img[@data-auto-id='product-GZ5554-01']")
                                select_color.click()
                            except:
                                pass
                        else:
                            try:
                                select_color = self.driver.find_element_by_xpath("//img[@data-auto-id='product-GW5350']")
                                select_color.click()
                            except:
                                pass
                            try:
                                select_color = self.driver.find_element_by_xpath("//img[@data-auto-id='product-GW5350-01']")
                                select_color.click()
                            except:
                                pass'''


                time.sleep(0.5)
                # потестить эти строчки на говно дропе
                # select_color = self.driver.find_element_by_xpath('//div[@data-auto-id='select-size']')
                # select_color.click()
                print(datetime.now())
                try:
                    select_box = WebDriverWait(self.driver, 60).until(
                        EC.element_to_be_clickable((By.CLASS_NAME, 'src-components-___sizes-dropdown__sizesDropdown___Dj9Z9  '))
                    )
                    #select_box = self.driver.find_element_by_class_name('src-components-___sizes-dropdown__sizesDropdown___Dj9Z9  ')
                    select_box.click()
                except:
                    print(232)
                    select_box1 = self.driver.find_elements(By.XPATH, '//div/div/div')
                    for x in select_box1:
                        if 'ВЫБРАТЬ РАЗМЕР' == x.text or 'Выбрать размер' == x.text or 'ВЫБРАТЬ Р' in x.text or 'Выбрать раз' in x.text:
                            x.click()
                            select_box = x
                            break

                options = self.driver.find_elements_by_xpath('//li[@data-auto-id="select-size-option"]')
                time.sleep(0.5)
                sizevector = []
                count = 19
                print(datetime.now())
                sizenumber = 0
                print(1)
                sizeselecter = len(options)//2 + (random.randint((-len(options)),(len(options))))//4
                try:
                    options[random.randint(0,len(options)-1)].click()
                finally:
                    time.sleep(0.2)
                '''for x in options:
                    if ('9 UK' in x.text or '42' in x.text or '9 UK' == x.text):
                        sizevector.append([2, sizenumber, x])
                    elif ('8.5 UK' in x.text or '41' in x.text or '8.5 UK' == x.text):
                        sizevector.append([1, sizenumber, x])
                    elif ('9.5 UK' in x.text or '42.5' in x.text or '9.5 UK' == x.text):
                        sizevector.append([3, sizenumber, x])
                    elif ('10 UK' in x.text or '43' in x.text or '10 UK' == x.text):
                        sizevector.append([4, sizenumber, x])
                    elif ('8 UK' in x.text or '40.5' in x.text or '8 UK' == x.text):
                        sizevector.append([5, sizenumber, x])
                    elif ('10.5 UK' in x.text or '44' in x.text or '10.5 UK' == x.text):
                        sizevector.append([8, sizenumber, x])
                    elif ('5.5 UK' in x.text or '37.5' in x.text or '5.5 UK' == x.text):
                        sizevector.append([2, sizenumber, x])
                    elif ('4.5 UK' in x.text or '36.5' in x.text or '4.5 UK' == x.text):
                        sizevector.append([9, sizenumber, x])
                    elif ('6 UK' in x.text or '38' in x.text or '6 UK' == x.text):
                        sizevector.append([10, sizenumber, x])
                    elif ('4 UK' in x.text or '36' in x.text or '4 UK' == x.text):
                        sizevector.append([11, sizenumber, x])
                    elif ('3.5 UK' in x.text or '35.5' in x.text or '3.5 UK' == x.text):
                        sizevector.append([12, sizenumber, x])
                    elif ('6.5 UK' in x.text or '38.5' in x.text or '6.5 UK' == x.text):
                        sizevector.append([13, sizenumber, x])
                    elif ('7 UK' in x.text or '39' in x.text or '7 UK' == x.text):
                        sizevector.append([14, sizenumber, x])
                    elif ('7.5 UK' in x.text or '40' in x.text or '7.5 UK' == x.text):
                        sizevector.append([15, sizenumber, x])
                    elif ('11 UK' in x.text or '44.5 ' in x.text or '11 UK' == x.text):
                        sizevector.append([16, sizenumber, x])
                    elif ('11.5 UK' in x.text or '45' in x.text or '11.5 UK' == x.text):
                        sizevector.append([17, sizenumber, x])
                    elif ('12 UK' in x.text or '46' in x.text or '12 UK' == x.text):
                        sizevector.append([18, sizenumber, x])
                    elif ('5 UK' in x.text or '37' in x.text or '5 UK' == x.text):
                        sizevector.append([6, sizenumber, x])
                    elif (' UK' in x.text):
                        sizevector.append([count, sizenumber, x])
                        count += 1
                    else:
                        sizevector.append([count, sizenumber, x])
                        count += 1
                    sizenumber += 1

                try:
                    sizevector.sort()
                except:
                    pass
                if sizenumber-1 < 8:
                    min8 = sizenumber-1
                else:
                    min8 = 8
                for h in range(random.randint(0, min8), 15):
                    try:
                        sizevector[h][2].click()
                        html = self.driver.page_source
                        if "К сожалению, выбранный размер закончился. Попробуйте заказать другой размер для примерки." in html:
                            select_box.click()
                        else:
                            break
                    except:
                        pass'''
                print(datetime.now())
                '''
                self.driver.get('https://www.adidas.ru/krossovki-stan-smith/FY5460.html')
                time.sleep(5)
                self.driver.find_element_by_xpath('/html/body/div[2]/div/div[1]/div/div/div/div[2]/div[2]/div[2]/section/div[1]/div[2]/button[7]/span').click()'''
                elem = self.driver.find_elements(By.XPATH, '//button/span')
                try:
                    for x in elem:
                        if (x.text == 'ДОБАВИТЬ В КОРЗИНУ'):
                            for i in range(1):
                                x.click()
                            break
                except:
                    pass
                '''time.sleep(1.5)
                try:
                    element = WebDriverWait(self.driver, 600).until(
                        EC.element_to_be_clickable((By.XPATH, '//span'))
                    )
                finally:
                    time.sleep(2)
                elem = self.driver.find_elements(By.XPATH, '//span')
                for x in elem:
                    if (x.text == 'Оfsd' or 'ПОСМ' in x.text or 'Посм' in x.text):
                        x.click()
                        break'''

                while True:
                    #self.driver.save_screenshot("page.png")
                    if (self.driver.current_url == 'https://www.adidas.ru/cart'):
                        break
                while True:
                    if ('1 товар' in self.driver.page_source or '1 Товар' in self.driver.page_source):
                        break
                try:
                    element = WebDriverWait(self.driver, 600).until(
                        EC.element_to_be_clickable((By.XPATH, '//span'))
                    )
                finally:
                    time.sleep(0.5)
                elem = self.driver.find_elements(By.XPATH, '//span')
                #self.driver.save_screenshot("page.png")
                try:
                    self.driver.find_element_by_xpath(
                        '/html/body/div[2]/div/div[1]/div/div/div/div[2]/div/aside/div[1]/div[1]/div/div[3]/div/button').click()
                except:
                    try:
                        actions = ActionChains(self.driver)
                        wait = WebDriverWait(self.driver, 20)
                        element = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR,
                                                                         ".cart-page__actions--aside___88R0p > div:nth-child(1) > button:nth-child(1) > span:nth-child(1)")))
                        actions.move_to_element(element).perform()
                        element.click()
                    except:
                        print(21)
                        for x in elem:
                            if (x.text == 'ОФОРМИТЬ' or 'Оформит' in x.text or 'ОФОРМИТ' in x.text):
                                x.click()
                                break
                while True:
                    if (self.driver.current_url == 'https://www.adidas.ru/delivery'):
                        break
                #self.driver.save_screenshot("page.png")
                try:
                    element = WebDriverWait(self.driver, 60).until(
                        EC.element_to_be_clickable((By.NAME, 'firstName'))
                    )
                finally:
                    time.sleep(0.5)
                whocoop = 5
                if(whocoop <= 10):
                    time.sleep(0.5)
                    firstName = self.driver.find_element_by_name('firstName')
                    #firstName.click()
                    firstName.send_keys(firstName1)
                    lastName = self.driver.find_element_by_name('lastName')
                    #lastName.click()
                    lastName.send_keys(lastName1)
                    city = self.driver.find_element_by_name('city')
                    #city.click()
                    city.send_keys(city1)
                    zipcode = self.driver.find_element_by_name('zipcode')
                    #zipcode.click()
                    zipcode.send_keys(zipcode1)
                    address1 = self.driver.find_element_by_name('address1')
                    #address1.click()
                    address1.send_keys(address11)
                    houseNumber = self.driver.find_element_by_name('houseNumber')
                    #houseNumber.click()
                    houseNumber.send_keys(houseNumber1)
                    phoneNumber = self.driver.find_element_by_name('phoneNumber')
                    #phoneNumber.click()
                    phoneNumber.send_keys(phoneNumber1)
                    emailAddress = self.driver.find_element_by_name('emailAddress')
                    #emailAddress.click()
                    emailAddress.send_keys(emailAddress1)
                    #self.driver.save_screenshot("pag1e.png")
                    try:
                        elem = self.driver.find_elements_by_class_name('gl-checkbox__input')
                        for x in elem:
                            x.click()
                    except:
                        elem = self.driver.find_element(By.XPATH,'//*[@id="app"]/div/div/div/div/div[2]/div/main/div[8]/div/div/label/input')
                        elem.click()
                    time.sleep(0.45)
                    try:
                        self.driver.find_element_by_xpath('/html/body/div[2]/div/div[1]/div/div/div/div[2]/div/main/div[10]/button/span').click()
                    except:
                        elem = self.driver.find_elements(By.XPATH, '//span')
                        for x in elem:
                            if (x.text == 'ПРОДОЛЖИТЬ ОФОРМЛЕНИЕ'):
                                x.click()
                                break
                    #self.driver.save_screenshot("page.png")
                    try:
                        element = WebDriverWait(self.driver, 60).until(
                            EC.element_to_be_clickable((By.CLASS_NAME, 'qa-iframe-widget'))
                        )
                    finally:
                        time.sleep(1)
                    iframe = self.driver.find_element_by_class_name('qa-iframe-widget')
                    self.driver.switch_to.frame(iframe)
                    elem = self.driver.find_element_by_xpath('//input[@name="card-number"]')
                    elem.click()
                    elem.clear()
                    elem.send_keys(str(sheet.cell(row=strokazapolnenia, column=13).value))
                    html = self.driver.page_source
                    while ("Проверьте номер карты" in html):
                        elem.send_keys(u'\ue009' + u'\ue003')
                        elem.clear()
                        elem.send_keys(str(sheet.cell(row=strokazapolnenia, column=13).value))
                        time.sleep(1)
                        html = self.driver.page_source
                    try:
                        elem = self.driver.find_element_by_name('expiry-name')
                        elem.send_keys(str(sheet.cell(row=strokazapolnenia, column=17)))
                    except:
                        elem = self.driver.find_element_by_name('expiry-month')
                        elem.click()
                        elem.send_keys(str(sheet.cell(row=strokazapolnenia, column=14).value))
                        elem = self.driver.find_element_by_name('expiry-year')
                        elem.click()
                        elem.send_keys(str(sheet.cell(row=strokazapolnenia, column=15).value))
                    elem = self.driver.find_element_by_name('security-code')
                    elem.click()
                    elem.send_keys(str(sheet.cell(row=strokazapolnenia, column=16).value))
                    time.sleep(0.5)
                    elem.send_keys(Keys.ENTER)

                    ds_sad(j, strokazapolnenia)
                    prevkod = ''
                    #self.driver.save_screenshot("page.png")
                    counterrrr -= 1
                    while True:
                        kod = ''
                        while (kod == '' or prevkod == kod):
                            try:
                                topic = str(sheet.cell(row=10, column=1).value)
                                urltopic = topic + '?offset=last'
                                response = requests.get(urltopic, headers={
                                    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:84.0) Gecko/20100101 Firefox/84.0"})
                                soup = BeautifulSoup(response.text, 'html.parser')
                                message = soup.find_all('div', {'class': 'bp_text'})
                                message = message[len(message) - 1].get_text()
                                splitted = message.split()
                                if (botnumber == splitted[0] and str(j) == splitted[1]):
                                    kod = splitted[2]
                            except:
                                pass

                        self.driver.switch_to.window(self.driver.window_handles[0])
                        try:
                            element = WebDriverWait(self.driver, 100).until(EC.presence_of_element_located((By.CLASS_NAME, 'qa-iframe-3ds-in-modal')))
                            iframe = self.driver.find_element_by_class_name('qa-iframe-3ds-in-modal')
                            self.driver.switch_to.frame(iframe)
                            try:
                                elem = self.driver.find_element_by_name('password')
                            except:
                                try:
                                    elem = self.driver.find_element_by_name('pwdInputVisible')
                                except:
                                    elem = self.driver.find_element_by_name('PASSWORD')
                            elem.click()
                            time.sleep(1)
                            elem.send_keys(str(kod))
                            try:
                                self.driver.find_element_by_xpath('//*[@id="btnSubmit"]').click()
                            except:
                                pass
                        except:
                            try:
                                elem = self.driver.find_element_by_name('PASSWORD')
                                elem.click()
                                elem.send_keys(str(kod))
                                elem.send_keys(Keys.ENTER)
                            except:
                                pass
                            try:
                                elem = self.driver.find_element_by_name('password')
                                elem.click()
                                time.sleep(1)
                                elem.send_keys(str(kod))
                            except:
                                pass
                            try:
                                elem = self.driver.find_element_by_name('DynamicPassword')
                                elem.click()
                                elem.send_keys(str(kod))
                                elem.send_keys(Keys.ENTER)
                            except:
                                pass
                        time.sleep(3)
                        print('Подтвердите платеж' in self.driver.page_source)
                        if('Подтвердите платеж' in self.driver.page_source or 'Неверный пароль' in self.driver.page_source or 'неверный код' in self.driver.page_source):
                            prevkod = kod
                            ds_sad1(j, strokazapolnenia)
                        else:
                            break
                else:
                    print(1)
            except:
                print(traceback.format_exc())
            finally:
                url = 'https://www.adidas.ru/cart'
                time.sleep(7)
                cardelem = self.driver.find_elements_by_class_name('gl-icon gl-cta__icon gl-cta__icon--standalone')
                for krestiki in cardelem:
                    krestiki.click()
                time.sleep(1)
    '''def c_f_m(self):
        a = {
            "type": "service_account",
            "project_id": "keen-oasis-309807",
            "private_key_id": "dada067581d935996b576fdb94d0bc86bf0fb145",
            "private_key": "-----BEGIN PRIVATE KEY-----\nMIIEvQIBADANBgkqhkiG9w0BAQEFAASCBKcwggSjAgEAAoIBAQCdqTAEnQUfxAAQ\nubWeYICuE3h/TkaWTuBzOxGxD4d+HE2hCve+GroBy5+gv3TmqnJznuxLf5+m6d/T\nnIcSfCdytLSt1s39FiHkbsJrrLOXKQYEdaudbjE3oC+zuUaPqEVxuTruFxgHW0zo\nT3668VB41+YI3AbFvmZieDZ3/AgLRihVzWkWQ32zADRjepTvleiDos+EEHBhU+yV\nuXu5EvDB/ReboFuPQpX+DVK8Pz5MquBpP3HfFv+pNseNAUl7zkFD+U1Zl5DdMWXk\nbDiQ+bJc/FLk0usZgyzM6vn9edtFd2uXn4kBmHAjTG07eJf0ttAZrxX2rJdOVoHo\n9Uw5JABnAgMBAAECggEAOAPW9CSnn+gez9JzZTBLN83iJ3Tznk6q7zXi7MIkNRmw\nORHxAIF1adLjRZUQXy3a2RpzQMW4vvBkHkaENPyYdWgfs43YhN16sDLr756d9gxU\nv3vhO99dLDLFD/k5SSwTdDrH5bSdzGtTUUyBLcE0AqNObwJhg5tIvB6S16FbaEl3\nfJDABwXch6CRy4xS7jOAQ3mJhISaNI/D/Nw2OWp0lQfbhOXkXA1Aw4dhRy4CdC0s\nj5n+Zx9P9pmsJTTmsL+LCGlJNlYib3FpfFROs5Ifjqg5CrvWWYHiQ1QVgdFz12Wq\n+YU0JP7LKqY7fXnLZq+w3Mk3cNGClODff5lCfC1MAQKBgQDSfYCf0S4Xbu1tjvHH\nG0Tgc5/qDyJozyuOFLiWXdJf+aME0YD+Z3di5T92NGStyHjV8iSiV0ExZY5byrPt\ntout1ljemvcEOg4ERa+iTZL+/6ALLcIm/lBaWyzSZo2bCRY84GFK3JgvRpXAr2fF\nVbFxrNO1+xYJEZCHCwyLo5nkAQKBgQC/v51lezivf/9YSQvwv1cDSo2rDyco40T8\ntP53GOpQcBVrBzEtc4F2QNmM+AE/OhpAdh7Aiz23/ZKF1tE6FmKtMx8s1OXnoFpQ\n2dQb9mXm8pAbETGYMjfjM/d7QLdFwFASMEuh1ZxfY4uDiRxOyuYVaVJELATnA2IF\nFiMtAalEZwKBgH1yEoAs2H1avOG4eVuiYtAp2LpTvrwcC+cnbkMfM0vDqWYrvebg\nnwQvgDzgSMhRW8KWB0NBITdhQRBvpAAd9OzrfBde36yDNOtS5LOj6+FRU36WDMnU\ncusGv5W9TNOC/N/XEeTdx00VkhY982hmA9StnZctRxqsfygca6OlIwgBAoGAS89/\ndFr2e3S4yPgAb4vG5Dc6sbIJAoJwiRUqMn44TdsyrKEoUVcEjXNwT+rR22SmcmwB\nFpE2sqLfpmjbSgMFuoKmxbvatexRKPwmf3O4sEmgrgLrP8hDDmJCa4mXrE0vmzFb\nBb4F2DTsOPIhrglMVJTSgL+CpK9Str4c5DNu7+sCgYEAos/6iFBcEo/NHHlHEke0\nepaaIp4TrV4wCCZN6KnxeDpCzE4qyuOZwO9zZexzxRBAbeRg+FBZ4GDYJEEWB6Cx\nm+27vOMuqgAmmGsxceb26nmE2PKqKfhZy/cQxftnbv9zvuUeK+HxGbJnFX3vRFXu\n8cI48iNjRiFprlzp8Fi8lck=\n-----END PRIVATE KEY-----\n",
            "client_email": "acc-833@keen-oasis-309807.iam.gserviceaccount.com",
            "client_id": "114178235896289023890",
            "auth_uri": "https://accounts.google.com/o/oauth2/auth",
            "token_uri": "https://oauth2.googleapis.com/token",
            "auth_provider_x509_cert_url": "https://www.googleapis.com/oauth2/v1/certs",
            "client_x509_cert_url": "https://www.googleapis.com/robot/v1/metadata/x509/acc-833%40keen-oasis-309807.iam.gserviceaccount.com"
        }
        credentials = ServiceAccountCredentials.from_json_keyfile_dict(a,
                                                                       ['https://www.googleapis.com/auth/spreadsheets',
                                                                        'https://www.googleapis.com/auth/drive'])
        gc = gspread.authorize(credentials)

        sh = gc.open_by_url('https://docs.google.com/spreadsheets/d/1ZUchdOpV6imkG-XYpiNKrM9SNzZMAoBbrmtjjgpSxBc')
        worksheet = sh.get_worksheet(0)
        values = worksheet.get_all_values()
        strokazapolnenia = random.randint(1, int(values[11][0]))
        firstName = self.driver.find_element_by_name('firstName')
        firstName.send_keys(str(values[1][4]))
        lastName = self.driver.find_element_by_name('lastName')
        lastName.send_keys(str(values[1][5]))
        city = self.driver.find_element_by_name('city')
        city.send_keys(str(values[1][6]))
        zipcode = self.driver.find_element_by_name('zipcode')
        zipcode.send_keys(str(values[1][7]))
        address1 = self.driver.find_element_by_name('address1')
        address1.send_keys(str(values[1][8]))
        houseNumber = self.driver.find_element_by_name('houseNumber')
        houseNumber.send_keys(str(values[1][9]))
        phoneNumber = self.driver.find_element_by_name('phoneNumber')
        phoneNumber.send_keys(str(values[1][10]))
        emailAddress = self.driver.find_element_by_name('emailAddress')
        emailAddress.send_keys(str(values[1][11]))
        try:
            elem = self.driver.find_elements_by_class_name('gl-checkbox__input')
            for x in elem:
                x.click()
        except:
            elem = self.driver.find_element(By.XPATH,
                                            '//*[@id="app"]/div/div/div/div/div[2]/div/main/div[8]/div/div/label/input')
            elem.click()
        time.sleep(0.45)
        try:
            self.driver.find_element_by_xpath(
                '/html/body/div[2]/div/div[1]/div/div/div/div[2]/div/main/div[10]/button/span').click()
        except:
            elem = self.driver.find_elements(By.XPATH, '//span')
            for x in elem:
                if (x.text == 'ПРОДОЛЖИТЬ ОФОРМЛЕНИЕ'):
                    x.click()
                    break
        try:
            element = WebDriverWait(self.driver, 60).until(
                EC.element_to_be_clickable((By.CLASS_NAME, 'qa-iframe-widget'))
            )
        finally:
            time.sleep(1)
        iframe = self.driver.find_element_by_class_name('qa-iframe-widget')
        self.driver.switch_to.frame(iframe)
        elem = self.driver.find_element_by_xpath('//input[@name="card-number"]')
        elem.click()
        elem.clear()
        elem.send_keys(str(values[1][12]))
        html = self.driver.page_source
        while ("Проверьте номер карты" in html):
            elem.send_keys(u'\ue009' + u'\ue003')
            elem.clear()
            elem.send_keys(str(values[1][12]))
            time.sleep(1)
            html = self.driver.page_source
        try:
            elem = self.driver.find_element_by_name('expiry-name')
            elem.send_keys(str(values[1][16]))
        except:
            elem = self.driver.find_element_by_name('expiry-month')
            elem.click()
            elem.send_keys(str(values[1][13]))
            elem = self.driver.find_element_by_name('expiry-year')
            elem.click()
            elem.send_keys(str(values[1][14]))
        elem = self.driver.find_element_by_name('security-code')
        elem.click()
        elem.send_keys(str(values[1][15]))
        time.sleep(1)
        elem.send_keys(Keys.ENTER)
        random1 = random.randint(100,10000)
        random2 = random.randint(100,10000)
        discord_webhook.DiscordWebhook(url=str(values[13][0]), content='Нужен код для ' + str(random1) + ' ' + str(random2) + ' вкладки ' + str(strokazapolnenia)).execute()
        prevkod = ''
        while True:
            kod = ''
            while (kod == '' or prevkod == kod):
                try:
                    topic = str(values[9][0])
                    urltopic = topic + '?offset=last'
                    response = requests.get(urltopic, headers={
                        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:84.0) Gecko/20100101 Firefox/84.0"})
                    soup = BeautifulSoup(response.text, 'html.parser')
                    message = soup.find_all('div', {'class': 'bp_text'})
                    message = message[len(message) - 1].get_text()
                    splitted = message.split()
                    if (str(random1) == splitted[0] and str(random2) == splitted[1]):
                        kod = splitted[2]
                except:
                    pass

            self.driver.switch_to.window(self.driver.window_handles[0])
            try:
                element = WebDriverWait(self.driver, 10).until(
                    EC.presence_of_element_located((By.CLASS_NAME, 'qa-iframe-3ds-in-modal')))
                iframe = self.driver.find_element_by_class_name('qa-iframe-3ds-in-modal')
                self.driver.switch_to.frame(iframe)
                try:
                    elem = self.driver.find_element_by_name('password')
                except:
                    try:
                        elem = self.driver.find_element_by_name('pwdInputVisible')
                    except:
                        elem = self.driver.find_element_by_name('PASSWORD')
                elem.click()
                time.sleep(1)
                elem.send_keys(str(kod))
                time.sleep(1)
                try:
                    self.driver.find_element_by_xpath('//*[@id="btnSubmit"]').click()
                except:
                    pass
            except:
                try:
                    elem = self.driver.find_element_by_name('PASSWORD')
                    elem.click()
                    elem.send_keys(str(kod))
                    elem.send_keys(Keys.ENTER)
                except:
                    pass
                try:
                    elem = self.driver.find_element_by_name('password')
                    elem.click()
                    time.sleep(1)
                    elem.send_keys(str(kod))
                except:
                    pass
                try:
                    elem = self.driver.find_element_by_name('DynamicPassword')
                    elem.click()
                    elem.send_keys(str(kod))
                    elem.send_keys(Keys.ENTER)
                except:
                    pass
            time.sleep(3)
            print('Подтвердите платеж' in self.driver.page_source)
            if (
                    'Подтвердите платеж' in self.driver.page_source or 'Неверный пароль' in self.driver.page_source or 'неверный код' in self.driver.page_source):
                prevkod = kod
                discord_webhook.DiscordWebhook(url=str(values[13][0]), content='ты бездарь Нужен код для ' + str(random1) + ' ' + str(random2) + ' вкладки ' + str(strokazapolnenia)).execute()
            else:
                break'''


def start():
    threads = []

    whichtask1 = 2

    whichtask2 = whichtask1 + int(sheet.cell(row=2, column=1).value)
    print(whichtask2)
    for whichtask in range(whichtask1, whichtask2, 1):
        try:

            run = adidas()
            hour = int(sheet.cell(row=whichtask, column=2).value)
            min = int(sheet.cell(row=whichtask, column=3).value)
            sec = int(sheet.cell(row=whichtask, column=4).value)
            thread = threading.Thread(target=(run.main), args=(hour, min, sec, whichtask - 1, sheet))
            threads.append(thread)
            thread.start()
            time.sleep(1)
        except:
            pass


if __name__ == '__main__':
    html = requests.get('https://bilet.pp.ru/calculator_rus/tochnoe_moskovskoe_vremia.php').text
    pattern = r">(.*?)-(.*?)-(.*?)<"
    today = re.findall(pattern, html)[0]
    print(int(today[0]))
    if int(today[2]) == 2021 and int(today[1]) <= 5 and int(today[0]) <= 30:
        print("OMR")
        start()
    else:
        print("ТЫ НЕ OMR :(")

#pyarmor pack -e " --onefile" --name "tempA" --clean py.py